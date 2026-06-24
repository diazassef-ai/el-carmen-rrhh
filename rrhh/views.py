import json
import calendar as calendar_lib
from collections import Counter, defaultdict
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import DocumentoAdjuntoForm, FuncionarioForm, PermisoCapacitacionForm, ReporteExcelForm
from .models import DocumentoAdjunto, Funcionario, PermisoCapacitacion, solicitudes_del_mes


PERFILES_RRHH = [
    "Administrador",
    "Secretaria",
    "Jefes de programa",
    "Solo Lectura",
    "Visualizacion calendario",
]
PERFILES_OPERATIVOS = [
    "Administrador",
    "Secretaria",
    "Jefes de programa",
    "Solo Lectura",
]
PERFILES_REPORTES = [
    "Administrador",
    "Secretaria",
    "Jefes de programa",
    "Solo Lectura",
]


def tiene_perfil(user, perfiles):
    return user.is_authenticated and user.groups.filter(name__in=perfiles).exists()


def es_admin(user):
    return user.is_superuser or user.groups.filter(name="Administrador").exists()


def puede_editar(user):
    return user.is_superuser or user.groups.filter(name__in=["Administrador", "Secretaria"]).exists()


def puede_ver_reportes(user):
    return user.is_authenticated and (
        user.is_superuser
        or tiene_perfil(user, PERFILES_REPORTES)
    )


def puede_ver_operativo(user):
    return user.is_authenticated and (user.is_superuser or tiene_perfil(user, PERFILES_OPERATIVOS))


def puede_ver_calendario(user):
    return user.is_authenticated and (user.is_superuser or tiene_perfil(user, PERFILES_RRHH))


def es_solo_calendario(user):
    return (
        user.is_authenticated
        and not user.is_superuser
        and user.groups.filter(name="Visualizacion calendario").exists()
        and not user.groups.filter(name__in=PERFILES_OPERATIVOS).exists()
    )


def asegurar_grupos_rrhh():
    for nombre in PERFILES_RRHH:
        Group.objects.get_or_create(name=nombre)


def aplicar_filtros_solicitudes(queryset, data):
    mes = data.get("mes")
    anio = data.get("anio")
    funcionario = data.get("funcionario")
    unidad = data.get("unidad")
    tipo_solicitud = data.get("tipo_solicitud")
    tipo_contrato = data.get("tipo_contrato")
    estado = data.get("estado")

    if mes:
        queryset = queryset.filter(fecha_desde__month=mes)
    if anio:
        queryset = queryset.filter(fecha_desde__year=anio)
    if funcionario:
        queryset = queryset.filter(funcionario=funcionario)
    if unidad:
        queryset = queryset.filter(funcionario__unidad__icontains=unidad)
    if tipo_solicitud:
        queryset = queryset.filter(tipo_solicitud=tipo_solicitud)
    if tipo_contrato:
        queryset = queryset.filter(funcionario__tipo_contrato=tipo_contrato)
    if estado:
        queryset = queryset.filter(estado=estado)
    return queryset


def construir_calendario_mensual(request):
    hoy = timezone.localdate()
    try:
        anio = int(request.GET.get("anio", hoy.year))
        mes = int(request.GET.get("mes", hoy.month))
        fecha_base = date(anio, mes, 1)
    except (TypeError, ValueError):
        fecha_base = date(hoy.year, hoy.month, 1)

    _, ultimo_dia = calendar_lib.monthrange(fecha_base.year, fecha_base.month)
    inicio_mes = fecha_base
    fin_mes = date(fecha_base.year, fecha_base.month, ultimo_dia)
    inicio_grilla = inicio_mes - timedelta(days=inicio_mes.weekday())
    fin_grilla = fin_mes + timedelta(days=6 - fin_mes.weekday())

    solicitudes = (
        PermisoCapacitacion.objects.select_related("funcionario")
        .filter(fecha_desde__lte=fin_grilla, fecha_hasta__gte=inicio_grilla)
        .exclude(estado__in=["Rechazado", "Anulado"])
        .order_by("fecha_desde", "funcionario__nombre_completo")
    )

    solicitudes_por_fecha = defaultdict(list)
    for solicitud in solicitudes:
        fecha = max(solicitud.fecha_desde, inicio_grilla)
        limite = min(solicitud.fecha_hasta, fin_grilla)
        while fecha <= limite:
            solicitudes_por_fecha[fecha].append(solicitud)
            fecha += timedelta(days=1)

    semanas = []
    fecha = inicio_grilla
    while fecha <= fin_grilla:
        semana = []
        for _ in range(7):
            semana.append({
                "fecha": fecha,
                "en_mes": fecha.month == fecha_base.month,
                "es_hoy": fecha == hoy,
                "solicitudes": solicitudes_por_fecha.get(fecha, []),
            })
            fecha += timedelta(days=1)
        semanas.append(semana)

    mes_anterior = fecha_base.replace(day=1) - timedelta(days=1)
    mes_siguiente = fin_mes + timedelta(days=1)
    meses = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    ]

    return {
        "semanas": semanas,
        "nombre_mes": f"{meses[fecha_base.month - 1].capitalize()} {fecha_base.year}",
        "mes_anterior": {"anio": mes_anterior.year, "mes": mes_anterior.month},
        "mes_siguiente": {"anio": mes_siguiente.year, "mes": mes_siguiente.month},
        "hoy": {"anio": hoy.year, "mes": hoy.month},
        "total_solicitudes_mes": solicitudes.filter(fecha_desde__lte=fin_mes, fecha_hasta__gte=inicio_mes).count(),
    }


@login_required
def dashboard(request):
    asegurar_grupos_rrhh()
    if es_solo_calendario(request.user):
        return redirect("rrhh_calendario_visualizacion")

    hoy = timezone.localdate()
    solicitudes_mes = solicitudes_del_mes(hoy)
    ausentes = PermisoCapacitacion.objects.vigentes_en(hoy)
    pendientes = PermisoCapacitacion.objects.filter(estado="Pendiente")
    aprobadas = PermisoCapacitacion.objects.filter(estado="Aprobado")

    por_unidad = list(
        aprobadas.values("funcionario__unidad")
        .annotate(total=Count("id"))
        .order_by("-total")[:8]
    )
    por_tipo = list(aprobadas.values("tipo_solicitud").annotate(total=Count("id")).order_by("-total"))
    por_mes = list(
        aprobadas.values("fecha_desde__year", "fecha_desde__month")
        .annotate(total=Count("id"))
        .order_by("fecha_desde__year", "fecha_desde__month")
    )
    top_funcionarios = list(
        aprobadas.values("funcionario__nombre_completo")
        .annotate(total_dias=Sum("numero_dias"), total=Count("id"))
        .order_by("-total_dias")[:8]
    )

    contexto = {
        "permisos_mes": solicitudes_mes.exclude(tipo_solicitud="Capacitacion").count(),
        "capacitaciones_mes": solicitudes_mes.filter(tipo_solicitud="Capacitacion").count(),
        "ausentes_hoy": ausentes.count(),
        "pendientes": pendientes.count(),
        "aprobadas": aprobadas.count(),
        "dias_acumulados": aprobadas.aggregate(total=Sum("numero_dias"))["total"] or 0,
        "por_unidad": json.dumps({
            "labels": [item["funcionario__unidad"] or "Sin unidad" for item in por_unidad],
            "values": [item["total"] for item in por_unidad],
        }),
        "por_tipo": json.dumps({
            "labels": [item["tipo_solicitud"] for item in por_tipo],
            "values": [item["total"] for item in por_tipo],
        }),
        "por_mes": json.dumps({
            "labels": [f"{item['fecha_desde__month']:02d}-{item['fecha_desde__year']}" for item in por_mes],
            "values": [item["total"] for item in por_mes],
        }),
        "top_funcionarios": top_funcionarios,
    }
    return render(request, "rrhh/dashboard.html", contexto)


@login_required
@user_passes_test(puede_ver_operativo, login_url="rrhh_dashboard")
def funcionarios_lista(request):
    q = request.GET.get("q", "").strip()
    funcionarios = Funcionario.objects.all()
    if q:
        funcionarios = funcionarios.filter(
            Q(nombre_completo__icontains=q)
            | Q(rut__icontains=q)
            | Q(cargo__icontains=q)
            | Q(unidad__icontains=q)
        )
    return render(request, "rrhh/funcionarios_lista.html", {"funcionarios": funcionarios, "q": q})


@login_required
@user_passes_test(puede_editar, login_url="rrhh_dashboard")
def funcionario_crear(request):
    form = FuncionarioForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        funcionario = form.save()
        messages.success(request, "Funcionario creado correctamente.")
        return redirect(funcionario)
    return render(request, "rrhh/formulario.html", {"form": form, "titulo": "Nuevo funcionario", "boton": "Guardar"})


@login_required
@user_passes_test(puede_ver_operativo, login_url="rrhh_dashboard")
def funcionario_ficha(request, pk):
    funcionario = get_object_or_404(Funcionario, pk=pk)
    solicitudes = funcionario.solicitudes.select_related("funcionario").prefetch_related("adjuntos")
    return render(request, "rrhh/funcionario_ficha.html", {"funcionario": funcionario, "solicitudes": solicitudes})


@login_required
@user_passes_test(puede_editar, login_url="rrhh_dashboard")
def funcionario_editar(request, pk):
    funcionario = get_object_or_404(Funcionario, pk=pk)
    form = FuncionarioForm(request.POST or None, instance=funcionario)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Ficha actualizada.")
        return redirect(funcionario)
    return render(request, "rrhh/formulario.html", {"form": form, "titulo": "Editar funcionario", "boton": "Guardar cambios"})


@login_required
@user_passes_test(puede_ver_operativo, login_url="rrhh_dashboard")
def solicitudes_lista(request):
    q = request.GET.get("q", "").strip()
    solicitudes = PermisoCapacitacion.objects.select_related("funcionario").prefetch_related("adjuntos")
    if q:
        solicitudes = solicitudes.filter(
            Q(funcionario__nombre_completo__icontains=q)
            | Q(funcionario__rut__icontains=q)
            | Q(funcionario__cargo__icontains=q)
            | Q(funcionario__unidad__icontains=q)
            | Q(tipo_solicitud__icontains=q)
        )
    return render(request, "rrhh/solicitudes_lista.html", {"solicitudes": solicitudes, "q": q})


@login_required
@user_passes_test(puede_editar, login_url="rrhh_dashboard")
def solicitud_crear(request):
    form = PermisoCapacitacionForm(request.POST or None)
    adjunto_form = DocumentoAdjuntoForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        solicitud = form.save()
        if request.FILES.get("archivo") and adjunto_form.is_valid():
            adjunto = adjunto_form.save(commit=False)
            adjunto.solicitud = solicitud
            adjunto.save()
        messages.success(request, "Solicitud registrada correctamente.")
        return redirect("rrhh_solicitud_detalle", pk=solicitud.pk)
    return render(request, "rrhh/solicitud_form.html", {"form": form, "adjunto_form": adjunto_form, "titulo": "Nueva solicitud"})


@login_required
@user_passes_test(puede_ver_operativo, login_url="rrhh_dashboard")
def solicitud_detalle(request, pk):
    solicitud = get_object_or_404(
        PermisoCapacitacion.objects.select_related("funcionario").prefetch_related("adjuntos"),
        pk=pk,
    )
    return render(request, "rrhh/solicitud_detalle.html", {"solicitud": solicitud})


@login_required
@user_passes_test(puede_editar, login_url="rrhh_dashboard")
def solicitud_editar(request, pk):
    solicitud = get_object_or_404(PermisoCapacitacion, pk=pk)
    form = PermisoCapacitacionForm(request.POST or None, instance=solicitud)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Solicitud actualizada.")
        return redirect("rrhh_solicitud_detalle", pk=solicitud.pk)
    return render(request, "rrhh/solicitud_form.html", {"form": form, "titulo": "Editar solicitud"})


@login_required
@user_passes_test(es_admin, login_url="rrhh_dashboard")
def solicitud_anular(request, pk):
    solicitud = get_object_or_404(PermisoCapacitacion, pk=pk)
    if request.method == "POST":
        solicitud.estado = "Anulado"
        solicitud.save(update_fields=["estado", "actualizado_en"])
        messages.warning(request, "Solicitud anulada.")
        return redirect("rrhh_solicitud_detalle", pk=solicitud.pk)
    return render(request, "rrhh/confirmar_anulacion.html", {"solicitud": solicitud})


@login_required
@user_passes_test(puede_editar, login_url="rrhh_dashboard")
def adjunto_crear(request, pk):
    solicitud = get_object_or_404(PermisoCapacitacion, pk=pk)
    form = DocumentoAdjuntoForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        adjunto = form.save(commit=False)
        adjunto.solicitud = solicitud
        adjunto.save()
        messages.success(request, "Documento adjunto guardado.")
        return redirect("rrhh_solicitud_detalle", pk=pk)
    return render(request, "rrhh/formulario.html", {"form": form, "titulo": "Adjuntar documento", "boton": "Subir"})


@login_required
@user_passes_test(puede_ver_calendario, login_url="rrhh_dashboard")
def calendario(request):
    return render(request, "rrhh/calendario.html")


@login_required
@user_passes_test(puede_ver_calendario, login_url="rrhh_dashboard")
def calendario_visualizacion(request):
    return render(
        request,
        "rrhh/calendario_visualizacion.html",
        construir_calendario_mensual(request),
    )


@login_required
@user_passes_test(puede_ver_calendario, login_url="rrhh_dashboard")
def calendario_eventos(request):
    solicitudes = PermisoCapacitacion.objects.select_related("funcionario").exclude(
        estado__in=["Rechazado", "Anulado"]
    )
    eventos = []
    for solicitud in solicitudes:
        eventos.append({
            "id": solicitud.pk,
            "title": f"{solicitud.funcionario.nombre_completo} - {solicitud.tipo_solicitud}",
            "start": solicitud.fecha_desde.isoformat(),
            "end": (solicitud.fecha_hasta + timedelta(days=1)).isoformat(),
            "color": solicitud.color_calendario,
            "url": request.build_absolute_uri(solicitud.get_absolute_url()) if hasattr(solicitud, "get_absolute_url") else "",
            "extendedProps": {
                "rut": solicitud.funcionario.rut,
                "cargo": solicitud.funcionario.cargo,
                "unidad": solicitud.funcionario.unidad,
                "tipo": solicitud.tipo_solicitud,
                "estado": solicitud.estado,
                "horario": solicitud.horario,
                "detalleUrl": request.build_absolute_uri(f"/rrhh/solicitudes/{solicitud.pk}/"),
            },
        })
    return JsonResponse(eventos, safe=False)


@login_required
@user_passes_test(puede_ver_operativo, login_url="rrhh_dashboard")
def ausentes_hoy(request):
    hoy = timezone.localdate()
    unidad = request.GET.get("unidad", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    ausentes = PermisoCapacitacion.objects.vigentes_en(hoy).select_related("funcionario")
    if unidad:
        ausentes = ausentes.filter(funcionario__unidad__icontains=unidad)
    if tipo:
        ausentes = ausentes.filter(tipo_solicitud=tipo)
    unidades = Funcionario.objects.order_by("unidad").values_list("unidad", flat=True).distinct()
    return render(request, "rrhh/ausentes_hoy.html", {
        "ausentes": ausentes,
        "unidades": unidades,
        "tipos": PermisoCapacitacion.TIPOS_SOLICITUD,
        "unidad": unidad,
        "tipo": tipo,
        "hoy": hoy,
    })


@login_required
@user_passes_test(puede_ver_reportes, login_url="rrhh_dashboard")
def reportes_excel(request):
    form = ReporteExcelForm(request.GET or None)
    solicitudes = PermisoCapacitacion.objects.select_related("funcionario").all()
    if form.is_valid():
        solicitudes = aplicar_filtros_solicitudes(solicitudes, form.cleaned_data)
    if request.GET.get("exportar") == "1" and form.is_valid():
        return construir_excel_rrhh(solicitudes)
    return render(request, "rrhh/reportes.html", {"form": form, "solicitudes": solicitudes[:100]})


def construir_excel_rrhh(solicitudes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen general"

    resumen_funcionarios = defaultdict(lambda: {"dias": 0, "permisos": 0, "capacitaciones": 0})
    dias_por_tipo = Counter()
    ranking = Counter()
    unidades = Counter()

    for solicitud in solicitudes:
        nombre = solicitud.funcionario.nombre_completo
        resumen_funcionarios[nombre]["dias"] += float(solicitud.numero_dias or 0)
        if solicitud.es_capacitacion:
            resumen_funcionarios[nombre]["capacitaciones"] += 1
        else:
            resumen_funcionarios[nombre]["permisos"] += 1
        dias_por_tipo[solicitud.tipo_solicitud] += float(solicitud.numero_dias or 0)
        ranking[nombre] += float(solicitud.numero_dias or 0)
        unidades[solicitud.funcionario.unidad] += 1

    escribir_tabla(ws, ["Funcionario", "Dias ausentes", "Permisos", "Capacitaciones"], [
        [nombre, datos["dias"], datos["permisos"], datos["capacitaciones"]]
        for nombre, datos in resumen_funcionarios.items()
    ])
    fila = ws.max_row + 3
    ws.cell(fila, 1, "Total de dias por tipo de ausencia").font = Font(bold=True)
    escribir_tabla(ws, ["Tipo de ausencia", "Dias"], dias_por_tipo.items(), fila + 1)

    ws_ranking = wb.create_sheet("Ranking")
    escribir_tabla(ws_ranking, ["Funcionario", "Dias ausentes"], ranking.most_common())

    ws_unidades = wb.create_sheet("Unidades")
    escribir_tabla(ws_unidades, ["Unidad", "Cantidad de ausencias"], unidades.items())

    ws_detalle = wb.create_sheet("Detalle completo")
    escribir_tabla(ws_detalle, [
        "Fecha inicio", "Fecha termino", "Nombre funcionario", "RUT", "Cargo",
        "Unidad", "Tipo de solicitud", "Numero de dias", "Estado", "Observaciones",
    ], [
        [
            s.fecha_desde, s.fecha_hasta, s.funcionario.nombre_completo, s.funcionario.rut,
            s.funcionario.cargo, s.funcionario.unidad, s.tipo_solicitud, s.numero_dias,
            s.estado, s.observaciones,
        ]
        for s in solicitudes
    ])

    for sheet in wb.worksheets:
        ajustar_hoja(sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_el_carmen_rrhh.xlsx"'
    wb.save(response)
    return response


def escribir_tabla(ws, encabezados, filas, fila_inicio=1):
    header_fill = PatternFill("solid", fgColor="0F766E")
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(fila_inicio, col, encabezado)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row_idx, fila in enumerate(filas, start=fila_inicio + 1):
        for col_idx, valor in enumerate(list(fila), start=1):
            ws.cell(row_idx, col_idx, valor)


def ajustar_hoja(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 14), 48)
    ws.freeze_panes = "A2"


@login_required
@user_passes_test(puede_ver_operativo, login_url="rrhh_dashboard")
def busqueda_global(request):
    q = request.GET.get("q", "").strip()
    funcionarios = Funcionario.objects.none()
    solicitudes = PermisoCapacitacion.objects.none()
    if q:
        funcionarios = Funcionario.objects.filter(
            Q(nombre_completo__icontains=q)
            | Q(rut__icontains=q)
            | Q(cargo__icontains=q)
            | Q(unidad__icontains=q)
        )
        solicitudes = PermisoCapacitacion.objects.select_related("funcionario").filter(
            Q(funcionario__nombre_completo__icontains=q)
            | Q(funcionario__rut__icontains=q)
            | Q(funcionario__cargo__icontains=q)
            | Q(funcionario__unidad__icontains=q)
            | Q(tipo_solicitud__icontains=q)
        )
    return render(request, "rrhh/busqueda.html", {"q": q, "funcionarios": funcionarios, "solicitudes": solicitudes})
